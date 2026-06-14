#!/usr/bin/env python3
"""创建竞品数据Excel文件"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = Workbook()
ws = wb.active
ws.title = "马来西亚-Insta360竞品"

# 表头
headers = ["排名", "产品ID", "产品名称", "店铺名称", "店铺类型", "商品类型", 
           "类目", "价格(RM)", "价格(¥)", "日销量", "周销量", "月销量",
           "日销售额", "周销售额", "月销售额", "累计销量", "评分", 
           "累计评论数", "累计点赞数", "上架时间"]

# 写入表头
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="2E75B6")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# 数据（从页面提取）
products = [
    [1, "26423668111", "Insta360 Flow 2 Pro Ultimate AI Tracking Stabilizer / Your Pocket AI Film Maker",
     "Insta360 Official Store", "本土", "虾皮商城", "Kamera & Drone>Aksesori Kamera>Gimbal & Penstabil",
     609, 1017.03, 2, 16, 65, 1218, 9744, 42885, 585, 4.99, 166, 847, "2025-01-16"],
    [2, "26886549144", "Insta360 Flow 2 Pro Foldable AI Phone Gimbal With Apple DockKit/Multi Tracking/360º Pan Tracking Built-In Tripod",
     "Insta360 Store", "跨境", "", "Kamera & Drone>Aksesori Kamera>Gimbal & Penstabil",
     759, 1267.53, 3, 12, 56, 2277, 8388, 36264, 946, 4.96, 223, 834, "2025-05-14"],
    [3, "42206775074", "Insta360 Flow 2 Your Pocket AI Filmmake Smartphone Gimbal Stabilizer With Built-In Selfie Stick / Tripod",
     "N4 Camera Store", "本土", "", "Kamera & Drone",
     359, 599.53, 0, 0, 43, 0, 0, 17117, 54, 5, 24, 51, "2025-07-01"],
    [4, "1398731782", "Insta360 Flow 2 Pro/Flow Pro/Flow - Al Powered Smartphone Gimbal Stabilizer With Built-In Selfie Stick",
     "DronesKaki", "本土", "", "Kamera & Drone>Aksesori Kamera>Gimbal & Penstabil",
     "398.8~758.8", "666~1267.2", 1, 8, 41, 399, 3190, 16351, 1000, 4.93, 527, 1229, "2018-08-09"],
    [5, "40205952961", "Insta360 Flow 2 Your Pocket AI Filmmaker",
     "Insta360 Official Store", "本土", "虾皮商城", "Kamera & Drone>Aksesori Kamera>Gimbal & Penstabil",
     389, 649.63, 0, 4, 22, 0, 1556, 9458, 195, 4.98, 51, 350, "2025-06-25"],
    [6, "28836645085", "Insta360 Flow 2 Pro Magnetic Phone Mount",
     "Insta360 Store", "跨境", "", "Telefon & Gajet>Aksesori>Pemegang Telefon",
     99, 165.33, 0, 3, 12, 0, 297, 1188, 320, 4.94, 90, 30, "2025-05-16"],
    [7, "41050932899", "Insta360 Flow 2 Pro USB-C Charge Cable",
     "Insta360 Store", "跨境", "", "Kamera & Drone>Lain-lain",
     39, 65.13, 0, 1, 3, 0, 39, 117, 41, 5, 13, 7, "2025-05-15"],
]

# 写入数据
for row_idx, product in enumerate(products, 2):
    for col_idx, value in enumerate(product, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        if col_idx == 1:  # 排名列居中
            cell.alignment = Alignment(horizontal="center")

# 设置列宽
col_widths = [6, 15, 80, 25, 10, 12, 55, 12, 12, 8, 8, 8, 12, 12, 12, 10, 8, 10, 10, 15]
for i, width in enumerate(col_widths, 1):
    ws.column_dimensions[chr(64 + i) if i <= 26 else f"A{chr(64 + i - 26)}"].width = width

# 保存
output_path = "/workspace/shopee-monitor/skills/shopee-scraper/data/竞品数据_MY_Insta360_20250614.xlsx"
wb.save(output_path)
print(f"Excel文件已保存: {output_path}")
